import tempfile
import unittest
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from ui.managers.file_manager import FileManager


class TestFileManager(unittest.TestCase):
    def test_save_xls_without_parent_falls_back_to_xlsx(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "legacy.xls"
            fallback_path = Path(tmpdir) / "legacy.xlsx"

            manager = FileManager()
            manager.excel_handler.df = pd.DataFrame({"index": ["01001"]})

            self.assertTrue(manager.save_file(str(path), parent=None))
            self.assertTrue(fallback_path.exists())
            self.assertEqual(manager.current_file, str(fallback_path))

    def test_save_xlsx_updates_current_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "saved.xlsx"

            manager = FileManager()
            manager.excel_handler.df = pd.DataFrame({"index": ["01001"]})

            self.assertTrue(manager.save_file(str(path), parent=None))
            self.assertTrue(path.exists())
            self.assertEqual(manager.current_file, str(path))

    def test_smart_save_preserves_unfiltered_rows_and_text_values(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "saved.xlsx"

            manager = FileManager()
            manager.excel_handler.original_df = pd.DataFrame({
                "city": ["Київ", "", "Львів"],
                "index": ["01001", "", "79000"],
                "_original_row_index": [0, 1, 2],
            })
            manager.excel_handler.df = pd.DataFrame({
                "_original_row_index": [0, 2],
                "city": ["Київ", "Львів"],
                "index": ["01002", "079000"],
            })

            self.assertTrue(manager.save_file(str(path), parent=None))
            reloaded = pd.read_excel(
                path,
                dtype=str,
                keep_default_na=False,
                na_filter=False,
                engine="openpyxl",
            )

        self.assertEqual(len(reloaded), 3)
        self.assertEqual(reloaded.loc[0, "index"], "01002")
        self.assertEqual(reloaded.loc[1, "index"], "")
        self.assertEqual(reloaded.loc[2, "index"], "079000")

    def test_smart_save_preserves_original_workbook_formatting_and_types(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / "source.xlsx"
            target_path = Path(tmpdir) / "target.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Аркуш2"
            worksheet.column_dimensions["A"].width = 18
            worksheet.column_dimensions["B"].width = 12
            worksheet.append(["ІПН", "Індекс", "Дата", "Сума"])
            worksheet.append(["30019905", "01001", datetime(2026, 5, 21), 20000])
            worksheet["C2"].number_format = "yyyy-mm-dd"
            workbook.save(source_path)

            manager = FileManager()
            handler = manager.excel_handler
            handler.load_file(str(source_path))
            handler.set_column_mapping({"index": [1]})
            handler.apply_column_filter()
            handler.df.iloc[0, handler.column_mapping["index"][0]] = "01002"
            manager.current_file = str(source_path)

            self.assertTrue(manager.save_file(str(target_path), parent=None))

            saved = load_workbook(target_path)
            saved_ws = saved.worksheets[0]

        self.assertEqual(saved_ws.title, "Аркуш2")
        self.assertEqual(saved_ws.column_dimensions["A"].width, 18)
        self.assertEqual(saved_ws["B2"].value, "01002")
        self.assertEqual(saved_ws["B2"].number_format, "@")
        self.assertIsInstance(saved_ws["C2"].value, datetime)
        self.assertEqual(saved_ws["C2"].number_format, "yyyy-mm-dd")
        self.assertEqual(saved_ws["D2"].value, 20000)


if __name__ == "__main__":
    unittest.main()
